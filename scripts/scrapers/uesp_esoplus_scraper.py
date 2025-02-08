import os
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# BASE_URL is used only for loading the list page.
BASE_URL = "https://en.uesp.net"

def get_esoplus_links():
    """
    Loads the ESO_Plus list page and returns a list of dictionaries.
    Each dictionary contains:
      - name: the ESO_Plus item's displayed name from the table.
      - webLink: the URL obtained from the item's name link.
    
    Processing rules for href:
      - If it starts with "//", prepend "https:".
      - If it starts with "http", leave it unchanged.
      - Otherwise, leave it unchanged.
    """
    url = BASE_URL + "/wiki/Online:Gallery_Furnishings/ESO_Plus"
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.content, "html.parser")
    
    esoplus_links = []
    table = soup.find("table", class_="wikitable")
    if table:
        rows = table.find_all("tr")
        # Skip the header row.
        for row in rows[1:]:
            cells = row.find_all("td")
            if not cells:
                continue
            # Assume the item's name link is in the second cell if available,
            # otherwise use the first cell.
            if len(cells) > 1:
                a_tag = cells[1].find("a")
            else:
                a_tag = cells[0].find("a")
            if a_tag and a_tag.has_attr("href"):
                href = a_tag["href"]
                if href.startswith("//"):
                    webLink = "https:" + href
                elif href.startswith("http"):
                    webLink = href
                else:
                    webLink = href
                name = a_tag.get_text(strip=True)
                esoplus_links.append({
                    "name": name,
                    "webLink": webLink
                })
    else:
        print("Could not find the ESO_Plus table on the page.")
    return esoplus_links

def get_raw_item_data(soup):
    """
    Extracts extra data from the detail page.
    
    It expects the extra data to be contained in:
      <div id="esoil_rawdata">
         <table id="esoil_rawdatatable">
            <tr>
              <td>itemId</td>
              <td id="">...</td>
            </tr>
            <tr>
              <td>allNames</td>
              <td id="">...</td>
            </tr>
            <tr>
              <td>description</td>
              <td id="">...</td>
            </tr>
            <tr>
              <td>icon</td>
              <td id="">...</td>
            </tr>
            <tr>
              <td>furnDataId</td>
              <td id="">...</td>
            </tr>
            <tr>
              <td>furnCategory</td>
              <td id="">...</td>
            </tr>
         </table>
      </div>
    
    Returns a dictionary of key/value pairs extracted from the table.
    """
    raw_data = {}
    div_data = soup.find("div", id="esoil_rawdata")
    if not div_data:
        return raw_data
    table = div_data.find("table", id="esoil_rawdatatable")
    if not table:
        return raw_data
    for row in table.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) >= 2:
            key = cells[0].get_text(separator=" ", strip=True)
            value = cells[1].get_text(separator=" ", strip=True)
            raw_data[key] = value
    return raw_data

def scrape_esoplus_data(item):
    """
    For a given ESO_Plus item (with 'name' and 'webLink' keys from the list page),
    loads the item's detail page via webLink, extracts the extra data from the
    <div id="esoil_rawdata">, and returns a dictionary with:
      - itemId
      - allNames
      - description
      - icon
      - furnDataId
      - furnCategory
      - webLink (from the list page)
      - name (from raw data if available; otherwise, the list page name)
    """
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(item["webLink"], headers=headers)
    soup = BeautifulSoup(response.content, "html.parser")
    
    raw_data = get_raw_item_data(soup)
    
    result = {
        "itemId": raw_data.get("itemId", ""),
        "allNames": raw_data.get("allNames", ""),
        "description": raw_data.get("description", ""),
        "icon": raw_data.get("icon", ""),
        "furnDataId": raw_data.get("furnDataId", ""),
        "furnCategory": raw_data.get("furnCategory", ""),
        "webLink": item["webLink"],
        "name": raw_data.get("name", item["name"])
    }
    return result

def export_to_excel(results, filename):
    """
    Exports the results to an Excel (.xlsx) file using openpyxl.
    Presets column widths for readability.
    """
    wb = Workbook()
    ws = wb.active
    headers = ["itemId", "webLink", "name", "allNames", "description", "icon", "furnDataId", "furnCategory"]
    ws.append(headers)
    for data in results:
        row = [data.get(col, "") for col in headers]
        ws.append(row)
    column_widths = {
        "itemId": 15,
        "webLink": 50,
        "name": 30,
        "allNames": 50,
        "description": 70,
        "icon": 30,
        "furnDataId": 20,
        "furnCategory": 20
    }
    for i, col in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = column_widths.get(col, 20)
    wb.save(filename)

def main():
    esoplus_items = get_esoplus_links()
    print(f"Found {len(esoplus_items)} ESO_Plus items on the list page.")
    
    results = []
    for item in esoplus_items:
        print(f"Scraping details for: {item['name']} ({item['webLink']})")
        data = scrape_esoplus_data(item)
        results.append(data)
        time.sleep(1)  # Pause for 1 second between requests.
    
    excel_filename = os.path.join("results", "esoplus_data.xlsx")
    export_to_excel(results, excel_filename)
    print(f"Export complete. Data saved to {excel_filename}")

if __name__ == "__main__":
    main()
