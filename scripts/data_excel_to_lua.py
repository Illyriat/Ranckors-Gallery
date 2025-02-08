import os
from openpyxl import load_workbook

def read_excel_data(filename):
    """
    Reads the given Excel file and returns a list of dictionaries.
    The first row is assumed to be headers.
    Expected headers (as produced by your scraper) are:
      itemId, webLink, name, allNames, description, icon, furnDataId, furnCategory
    """
    wb = load_workbook(filename)
    ws = wb.active
    data = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
        else:
            entry = {}
            for j, cell in enumerate(row):
                header = headers[j]
                entry[header] = cell if cell is not None else ""
            data.append(entry)
    return data

def format_value(key, value):
    """
    Formats a value for Lua output.
    For 'itemId', if the value is numeric, it is output without quotes.
    For all other keys, the value is output as a quoted string with inner quotes escaped.
    """
    if key == "itemId":
        try:
            num = int(float(value))
            return str(num)
        except (ValueError, TypeError):
            pass
    if not isinstance(value, str):
        value = str(value)
    escaped = value.replace('"', '\\"')
    return f'"{escaped}"'

def convert_to_lua(data, table_name):
    """
    Converts the list of dictionaries (Excel rows) into a Lua table formatted string.
    The Excel column "webLink" is mapped to the Lua key "link".
    The resulting Lua table is assigned to the variable named table_name.
    """
    lua_lines = []
    lua_lines.append(f"local {table_name} = {{")
    for entry in data:
        lua_lines.append("    {")
        lua_lines.append(f'        icon = {format_value("icon", entry.get("icon", ""))},')
        lua_lines.append(f'        itemId = {format_value("itemId", entry.get("itemId", ""))},')
        lua_lines.append(f'        link = {format_value("link", entry.get("webLink", ""))},')
        lua_lines.append(f'        name = {format_value("name", entry.get("name", ""))},')
        lua_lines.append(f'        allNames = {format_value("allNames", entry.get("allNames", ""))},')
        lua_lines.append(f'        furnDataId = {format_value("furnDataId", entry.get("furnDataId", ""))},')
        lua_lines.append(f'        furnCategory = {format_value("furnCategory", entry.get("furnCategory", ""))},')
        lua_lines.append(f'        description = {format_value("description", entry.get("description", ""))},')
        lua_lines.append("    },")
    lua_lines.append("}")
    return "\n".join(lua_lines)

def main():
    # Define paths to the Excel files (assumed to be in the "results" folder).
    paintings_excel   = os.path.join("results", "paintings_data.xlsx")
    music_boxes_excel = os.path.join("results", "music_boxes_data.xlsx")
    banners_excel     = os.path.join("results", "banners_data.xlsx")
    tapestries_excel  = os.path.join("results", "tapestries_data.xlsx")
    esoplus_excel     = os.path.join("results", "esoplus_data.xlsx")
    literature_excel  = os.path.join("results", "literature_data.xlsx")
    maps_excel        = os.path.join("results", "maps_data.xlsx")
    
    # Define output file paths in the data folder (one level up).
    paintings_lua_file   = os.path.join("..", "data", "paintings.lua")
    music_boxes_lua_file = os.path.join("..", "data", "music_box.lua")
    banners_lua_file     = os.path.join("..", "data", "banners.lua")
    tapestries_lua_file  = os.path.join("..", "data", "tapestries.lua")
    esoplus_lua_file     = os.path.join("..", "data", "esoplus.lua")
    literature_lua_file  = os.path.join("..", "data", "literature.lua")
    maps_lua_file        = os.path.join("..", "data", "maps.lua")
    
    # Process Paintings
    print("Processing paintings...")
    paintings_data = read_excel_data(paintings_excel)
    paintings_lua = convert_to_lua(paintings_data, "paintings")
    with open(paintings_lua_file, "w", encoding="utf-8") as f:
        f.write(paintings_lua)
    print(f"Lua file created for paintings: {paintings_lua_file}")
    
    # Process Music Boxes
    print("Processing music boxes...")
    music_boxes_data = read_excel_data(music_boxes_excel)
    music_boxes_lua = convert_to_lua(music_boxes_data, "musicBoxes")
    with open(music_boxes_lua_file, "w", encoding="utf-8") as f:
        f.write(music_boxes_lua)
    print(f"Lua file created for music boxes: {music_boxes_lua_file}")
    
    # Process Banners
    print("Processing banners...")
    banners_data = read_excel_data(banners_excel)
    banners_lua = convert_to_lua(banners_data, "banners")
    with open(banners_lua_file, "w", encoding="utf-8") as f:
        f.write(banners_lua)
    print(f"Lua file created for banners: {banners_lua_file}")
    
    # Process Tapestries
    print("Processing tapestries...")
    tapestries_data = read_excel_data(tapestries_excel)
    tapestries_lua = convert_to_lua(tapestries_data, "tapestries")
    with open(tapestries_lua_file, "w", encoding="utf-8") as f:
        f.write(tapestries_lua)
    print(f"Lua file created for tapestries: {tapestries_lua_file}")
    
    # Process ESO_Plus
    print("Processing ESO_Plus...")
    esoplus_data = read_excel_data(esoplus_excel)
    esoplus_lua = convert_to_lua(esoplus_data, "esoplus")
    with open(esoplus_lua_file, "w", encoding="utf-8") as f:
        f.write(esoplus_lua)
    print(f"Lua file created for ESO_Plus: {esoplus_lua_file}")
    
    # Process Literature
    print("Processing Literature...")
    literature_data = read_excel_data(literature_excel)
    literature_lua = convert_to_lua(literature_data, "literature")
    with open(literature_lua_file, "w", encoding="utf-8") as f:
        f.write(literature_lua)
    print(f"Lua file created for Literature: {literature_lua_file}")
    
    # Process Maps
    print("Processing Maps...")
    maps_data = read_excel_data(maps_excel)
    maps_lua = convert_to_lua(maps_data, "maps")
    with open(maps_lua_file, "w", encoding="utf-8") as f:
        f.write(maps_lua)
    print(f"Lua file created for Maps: {maps_lua_file}")

if __name__ == "__main__":
    main()
