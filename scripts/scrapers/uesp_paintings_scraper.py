import os
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# BASE_URL is used only to load the list page.
BASE_URL = "https://en.uesp.net"

def get_painting_links():
    """
    Loads the UESP paintings list page and returns a list of dictionaries.
    Each dictionary contains:
      - name: the painting's displayed name from the table.
      - webLink: the URL obtained from the painting's name link.
    
    Processing rules for href:
      - If it starts with "//", prepend "https:".
      - If it starts with "http", leave it unchanged.
      - Otherwise, leave it unchanged.
    """
    url = BASE_URL + "/wiki/Online:Gallery_Furnishings/Paintings"
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.content, "html.parser")
    
    painting_links = []
    table = soup.find("table", class_="wikitable")
    if table:
        rows = table.find_all("tr")
        # Skip the header row.
        for row in rows[1:]:
            cells = row.find_all("td")
            if not cells:
                continue
            # Assume the painting name (with its link) is in the second cell.
            if len(cells) > 1:
                a_tag = cells[1].find("a")
            else:
                a_tag = cells[0].find("a")
            if a_tag and a_tag.has_attr("href"):
                href = a_tag["href"]
                # Process href without appending any BASE_URL.
                if href.startswith("//"):
                    webLink = "https:" + href
                elif href.startswith("http"):
                    webLink = href
                else:
                    webLink = href
                name = a_tag.get_text(strip=True)
                painting_links.append({
                    "name": name,
                    "webLink": webLink
                })
    else:
        print("Could not find the paintings table on the page.")
    return painting_links

def get_raw_item_data(soup):
    """
    Extracts extra data from the detail page.
    
    Expects the extra data to be contained in:
      <div id="esoil_rawdata">
         <table id="esoil_rawdatatable">
            <tr>
              <td>itemId</td>
              <td id="">204807</td>
            </tr>
            <tr>
              <td>allNames</td>
              <td id="">A Clear Day in Colovia Painting, Metal; ...</td>
            </tr>
            <tr>
              <td>description</td>
              <td id="">[Some description text]</td>
            </tr>
            <tr>
              <td>icon</td>
              <td id="">[Icon string]</td>
            </tr>
            <tr>
              <td>furnDataId</td>
              <td id="">[FurnDataId value]</td>
            </tr>
            <tr>
              <td>furnCategory</td>
              <td id="">[FurnCategory value]</td>
            </tr>
         </table>
      </div>
    
    Returns a dictionary of key/value pairs extracted from the table.
    """
    raw_data = {}
    div_data = soup.find("div", id="esoil_rawdata")
    if not div_data:
        return raw_data
    table = div_data.find("table", id="esoil_rawdatatable")
    if not table:
        return raw_data
    for row in table.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) >= 2:
            key = cells[0].get_text(separator=" ", strip=True)
            value = cells[1].get_text(separator=" ", strip=True)
            raw_data[key] = value
    return raw_data

def scrape_painting_data(painting):
    """
    For a given painting (with 'name' and 'webLink' keys from the list page),
    loads the painting detail page via webLink, extracts the extra data from
    the <div id="esoil_rawdata">, and returns a dictionary with:
      - itemId
      - allNames
      - description
      - icon
      - furnDataId
      - furnCategory
      - webLink (from the list page)
      - name (from raw data if available; otherwise, the list page name)
    """
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(painting["webLink"], headers=headers)
    soup = BeautifulSoup(response.content, "html.parser")
    
    raw_data = get_raw_item_data(soup)
    
    result = {
        "itemId": raw_data.get("itemId", ""),
        "allNames": raw_data.get("allNames", ""),
        "description": raw_data.get("description", ""),
        "icon": raw_data.get("icon", ""),
        "furnDataId": raw_data.get("furnDataId", ""),
        "furnCategory": raw_data.get("furnCategory", ""),
        "webLink": painting["webLink"],
        "name": raw_data.get("name", painting["name"]),
        "link": raw_data.get("link", "")
    }
    return result

def export_to_excel(results, filename):
    """
    Exports the results to an Excel (.xlsx) file using openpyxl.
    Sets preset column widths for readability.
    """
    wb = Workbook()
    ws = wb.active
    headers = ["itemId", "webLink", "name", "allNames", "description", "icon", "furnDataId", "furnCategory", "link"]
    ws.append(headers)
    
    # Append each data row.
    for data in results:
        row = [data.get(col, "") for col in headers]
        ws.append(row)
    
    # Set preset column widths.
    column_widths = {
        "itemId": 15,
        "webLink": 50,
        "name": 30,
        "allNames": 50,
        "description": 70,
        "icon": 30,
        "furnDataId": 20,
        "furnCategory": 20,
        "link": 40
    }
    for i, col in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = column_widths.get(col, 20)
    
    wb.save(filename)

def main():
    # Create the results folder if it doesn't exist.
    results_folder = "results"
    if not os.path.exists(results_folder):
        os.makedirs(results_folder)
    
    paintings = get_painting_links()
    print(f"Found {len(paintings)} paintings on the list page.")
    
    results = []
    for painting in paintings:
        print(f"Scraping details for: {painting['name']} ({painting['webLink']})")
        data = scrape_painting_data(painting)
        results.append(data)
        # Pause for 1 second between requests to be respectful.
        time.sleep(1)
    
    excel_filename = os.path.join(results_folder, "paintings_data.xlsx")
    export_to_excel(results, excel_filename)
    print(f"Export complete. Data saved to {excel_filename}")

if __name__ == "__main__":
    main()
